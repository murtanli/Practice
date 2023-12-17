import openpyxl as ex
import string
from accessify import protected
from openpyxl.styles import Alignment

class sortir_excel:
    def __init__(self, path):
        self.path = path
        self.mas_res = []
        self.alphabet = list(string.ascii_lowercase)

    def open_excel(self):
        document = ex.open(str(self.path))
        sheet = document.active
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            if row[0].value is None:
                break
            currect_mas = [[]]
            for i in row:
                try:
                    currect_mas.append(int(i.value))
                except ValueError:
                    currect_mas[0].append(self.alphabet.index(i.value))
            self.mas_res.append(currect_mas)
        document.close()
        return self.sort_mas()


    @protected
    def sort_mas(self):
        amount_mas_val = len(self.mas_res)
        for i in range(amount_mas_val):
            for j in range(amount_mas_val - i - 1):
                if self.mas_res[j][0][0] > self.mas_res[j + 1][0][0]:
                    self.mas_res[j], self.mas_res[j + 1] = self.mas_res[j + 1], self.mas_res[j]
                elif self.mas_res[j][0][0] == self.mas_res[j + 1][0][0] and self.mas_res[j][0][1] > self.mas_res[j + 1][0][1]:
                    self.mas_res[j], self.mas_res[j + 1] = self.mas_res[j + 1], self.mas_res[j]
        return self.summ()

    @protected
    def summ(self):
        indices = []
        amount_mas_val = range(len(self.mas_res) - 1)
        for i in amount_mas_val:
            if self.mas_res[i][0][0] == self.mas_res[i + 1][0][0] and self.mas_res[i][0][1] == self.mas_res[i + 1][0][1]:
                self.mas_res[i + 1][1] += self.mas_res[i][1]
            else:
                indices.append(self.mas_res[i])
        self.mas_res = indices
        return self.write_into_excel()

    @protected
    def write_into_excel(self):
        try:
            document = ex.open(str(self.path))
            sheet = document.active
            sheet[f'G{1}'] = 'String'
            sheet[f'H{1}'] = 'String'
            sheet[f'I{1}'] = 'Integer'
            for i in range(len(self.mas_res)):
                sheet[f'F{i + 2}'] = i + 1
                sheet[f'G{i + 2}'] = self.alphabet[self.mas_res[i][0][0]]
                sheet[f'H{i + 2}'] = self.alphabet[self.mas_res[i][0][1]]
                sheet[f'I{i + 2}'] = self.mas_res[i][1]
                sheet[f'F{i + 2}'].alignment = Alignment(horizontal='center', vertical='center')
                sheet[f'G{i + 2}'].alignment = Alignment(horizontal='center', vertical='center')
                sheet[f'H{i + 2}'].alignment = Alignment(horizontal='center', vertical='center')
                sheet[f'I{i + 2}'].alignment = Alignment(horizontal='center', vertical='center')
                #print(self.mas_res[i])
            document.save("New_file.xlsx")
        except PermissionError:
            return("Закрой файл !! - New_file.xlsx")
        else:
            return("Выполение закончено, - New_file.xlsx")


sr = sortir_excel("document.xlsx")
print(sr.open_excel())
